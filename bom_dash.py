import re, io
import pandas as pd
import dash
from dash import dcc, html, dash_table, Input, Output, State
import dash_bootstrap_components as dbc
from openpyxl import load_workbook

# ─── CONFIG ────────────────────────────────────────────────────────────────────
import os
INPUT_PATH = os.path.join(os.path.dirname(__file__), "plan_produccion.xlsx")
STOCK_ALMS = ["ALMA002", "ALMA089", "ALMA070", "ALMA071", "ALMA072"]

# ─── CARGA ──────────────────────────────────────────────────────────────────────
def load_data(path):
    def find_sheet(wb_names, keywords):
        for n in wb_names:
            if any(k.lower() in n.lower() for k in keywords):
                return n
        raise ValueError(f"Hoja no encontrada: {keywords}. Disponibles: {wb_names}")

    # BOM — leer con openpyxl para evitar que pandas interprete "5,500" como 5.5
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    bom_sheet = find_sheet(sheet_names, ["bom"])
    sh = wb[bom_sheet]
    col_names = ["cod_pt","posicion","desc_pt","und_pt","qty_pt","ceco",
                 "cod_comp","desc_comp","und_comp","qty_comp"]
    rows = []
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if i == 0:
            continue
        r = list(row) + [None] * 10
        rows.append(r[:10])
    wb.close()

    df_bom = pd.DataFrame(rows, columns=col_names)
    df_bom = df_bom[df_bom["cod_pt"].notna() & df_bom["cod_comp"].notna()].copy()
    df_bom["cod_pt"]   = df_bom["cod_pt"].astype(str).str.strip()
    df_bom["cod_comp"] = df_bom["cod_comp"].astype(str).str.strip()
    df_bom = df_bom[df_bom["cod_pt"].str.match(r"^\d+$")].copy()
    for col in ["qty_pt","qty_comp"]:
        df_bom[col] = pd.to_numeric(df_bom[col], errors="coerce").fillna(0)

    # Stock
    xl = pd.ExcelFile(path)
    df_stock = xl.parse(find_sheet(sheet_names, ["stock"]), dtype=str)
    scols = {df_stock.columns[0]:"cod", df_stock.columns[2]:"alm_cod",
             df_stock.columns[4]:"en_stock", df_stock.columns[5]:"um",
             df_stock.columns[8]:"precio_prom"}
    df_stock = df_stock.rename(columns=scols)
    df_stock["cod"]         = df_stock["cod"].astype(str).str.strip()
    df_stock["en_stock"]    = pd.to_numeric(df_stock["en_stock"].astype(str).str.replace(",",""), errors="coerce").fillna(0)
    df_stock["precio_prom"] = pd.to_numeric(df_stock["precio_prom"].astype(str).str.replace(",",""), errors="coerce").fillna(0)
    df_stock = df_stock[df_stock["alm_cod"].str.upper().isin([a.upper() for a in STOCK_ALMS])]
    df_stock = df_stock.groupby("cod", as_index=False).agg(
        en_stock=("en_stock","sum"), precio_prom=("precio_prom","mean"), um=("um","first"))

    # Spec
    df_spec = xl.parse(find_sheet(sheet_names, ["spec","bd_spec"]), dtype=str)
    sp = {df_spec.columns[0]:"cod", df_spec.columns[3]:"tipo_mat", df_spec.columns[13]:"tipo_mat2"}
    df_spec = df_spec.rename(columns=sp)
    df_spec["cod"] = df_spec["cod"].astype(str).str.strip()
    df_spec = df_spec[["cod","tipo_mat","tipo_mat2"]].drop_duplicates("cod")

    # Tiempos — cantidad base y tiempos de producción de semiterminados
    try:
        df_tiempos = xl.parse(find_sheet(sheet_names, ["tiempo","tiempos"]), dtype=str)
        # columnas: Código Semi, desc, Cantidad Base, T.MO, T.Maq, Cant.Opr, Maquina, ...
        tc = {df_tiempos.columns[0]:"cod_semi", df_tiempos.columns[1]:"desc_semi",
              df_tiempos.columns[2]:"cant_base", df_tiempos.columns[3]:"t_mo",
              df_tiempos.columns[4]:"t_maq",     df_tiempos.columns[5]:"cant_opr",
              df_tiempos.columns[6]:"maquina",   df_tiempos.columns[9]:"proceso"}
        df_tiempos = df_tiempos.rename(columns=tc)
        df_tiempos["cod_semi"] = df_tiempos["cod_semi"].astype(str).str.strip()
        for col in ["cant_base","t_mo","t_maq","cant_opr"]:
            df_tiempos[col] = pd.to_numeric(
                df_tiempos[col].astype(str).str.replace(",",""), errors="coerce").fillna(0)
        df_tiempos = df_tiempos[df_tiempos["cod_semi"].str.match(r"^\d+$")].copy()
    except Exception as e:
        print(f"  Tiempos no cargados: {e}")
        df_tiempos = pd.DataFrame(columns=["cod_semi","desc_semi","cant_base","t_mo","t_maq","cant_opr","maquina","proceso"])

    # Ordenes — MOQ, LT-días, Tipo de compra para materiales comprados
    try:
        df_ord = xl.parse("Materiales", dtype=str)
        oc = {df_ord.columns[0]:"cod",
              df_ord.columns[3]:"tipo_compra",
              df_ord.columns[4]:"moq",
              df_ord.columns[5]:"lt_dias"}
        df_ord = df_ord.rename(columns=oc)
        df_ord["cod"]       = df_ord["cod"].astype(str).str.strip()
        df_ord["tipo_compra"] = df_ord["tipo_compra"].astype(str).str.strip()
        df_ord["moq"]       = pd.to_numeric(df_ord["moq"].astype(str).str.replace(",",""), errors="coerce").fillna(0)
        df_ord["lt_dias"]   = pd.to_numeric(df_ord["lt_dias"].astype(str).str.replace(",",""), errors="coerce")
        # NO rellenar con 0 — dejar NaN para distinguir "sin LT definido" de "LT=0"
        df_ord = df_ord[["cod","tipo_compra","moq","lt_dias"]].drop_duplicates("cod")
        # filtrar solo los que tienen código numérico válido
        df_ord = df_ord[df_ord["cod"].str.match(r"^\d+$")].copy()
        n_con_lt = df_ord["lt_dias"].notna().sum()
        print(f"  Materiales: {len(df_ord)} registros | {n_con_lt} con LT definido | muestra: {df_ord[df_ord['lt_dias'].notna()][['cod','lt_dias']].head(3).to_dict('records')}")
    except Exception as e:
        print(f"  Materiales no cargados: {e}")
        df_ord = pd.DataFrame(columns=["cod","tipo_compra","moq","lt_dias"])

    return df_bom, df_stock, df_spec, df_tiempos, df_ord

def classify(cod, cod_raiz=""):
    c = str(cod).strip()
    r = str(cod_raiz).strip()
    if c.startswith("211"): return "PT Útiles"
    if c.startswith("214"): return "PT Cuadernos"
    if c.startswith("231"): return "Semiterminado"
    # 232x → Semiterminado solo si el PT raíz es cuadernos (214011)
    if c.startswith("232") and r.startswith("214011"): return "Semiterminado"
    return "Comprado"

# ─── EXPLOSIÓN ──────────────────────────────────────────────────────────────────
def explode_bom(df_bom, pt_qty_map, df_stock, df_spec, pt_desc=None):
    # qty_pt por código padre
    if pt_desc is None: pt_desc = {}
    qty_pt_map = df_bom.groupby("cod_pt")["qty_pt"].first().to_dict()
    bom_dict   = {}
    for _, row in df_bom.iterrows():
        bom_dict.setdefault(row["cod_pt"], []).append(row)

    records = []

    def recurse(cod_raiz, parent_cod, qty_acum, level):
        if parent_cod not in bom_dict:
            return
        qty_pt_base = qty_pt_map.get(parent_cod, 1.0) or 1.0
        for row in bom_dict[parent_cod]:
            comp      = row["cod_comp"]
            qty_unit  = row["qty_comp"] / qty_pt_base   # consumo por 1 unidad de PT
            qty_total = qty_acum * qty_unit
            tipo      = classify(comp, cod_raiz)

            stk = df_stock[df_stock["cod"] == comp]
            en_stock    = float(stk["en_stock"].values[0])    if len(stk) else 0.0
            precio_prom = float(stk["precio_prom"].values[0]) if len(stk) else 0.0
            um          = stk["um"].values[0]                  if len(stk) else str(row["und_comp"] or "")

            sp = df_spec[df_spec["cod"] == comp]
            tipo_mat  = sp["tipo_mat"].values[0]  if len(sp) else ""
            tipo_mat2 = sp["tipo_mat2"].values[0] if len(sp) else ""

            necesidad  = round(qty_total, 6)
            diferencia = round(en_stock - necesidad, 6)

            records.append({
                "PT_Raiz":         cod_raiz,
                "Desc_PT_Raiz":    str(pt_desc.get(cod_raiz, "")),
                "Nivel":           level,
                "Tipo":            tipo,
                "Posicion":        row["posicion"],
                "Cod_Componente":  comp,
                "Desc_Componente": str(row["desc_comp"] or ""),
                "Und":             um,
                "Qty_BOM_Base":    round(row["qty_pt"], 2),
                "Qty_Comp_BOM":    round(row["qty_comp"], 4),
                "Qty_Unit":        round(qty_unit, 6),
                "Qty_Necesaria":   necesidad,
                "En_Stock":        round(en_stock, 3),
                "Diferencia":      diferencia,
                "Precio_Prom_S/":  round(precio_prom, 4),
                "Valor_Nec_S/":    round(necesidad * precio_prom, 2),
                "Tipo_Mat":        str(tipo_mat or ""),
                "Tipo_Mat2":       str(tipo_mat2 or ""),
                "CECO":            str(row["ceco"] or ""),
                "PT_Padre":        parent_cod,
            })

            if tipo == "Semiterminado":
                recurse(cod_raiz, comp, qty_total, level + 1)

    for pt, qty_pedida in pt_qty_map.items():
        if pt not in bom_dict:
            continue
        recurse(pt, pt, float(qty_pedida), 1)

    return pd.DataFrame(records) if records else pd.DataFrame()

# ─── NETOS: explosión descontando stock de semiterminados ───────────────────────
def calcular_netos(df_bom, pt_qty_map, df_stock, df_spec):
    """
    Lógica:
    - Para cada semiterminado (231x) en la explosión, verifica si hay stock disponible.
    - Si stock >= necesidad → NO se explota ese semiterminado (se usa el stock).
    - Si stock < necesidad → se explota solo la diferencia (necesidad - stock).
    - El resultado final son solo los materiales que realmente hay que COMPRAR o PRODUCIR.
    """
    qty_pt_map = df_bom.groupby("cod_pt")["qty_pt"].first().to_dict()
    bom_dict   = {}
    for _, row in df_bom.iterrows():
        bom_dict.setdefault(row["cod_pt"], []).append(row)

    # Stock inicial por código (se reinicia por PT raíz para no contaminar entre PTs)
    stock_inicial = {}
    stk_rows = df_stock.set_index("cod")
    for cod in stk_rows.index:
        stock_inicial[cod] = float(stk_rows.loc[cod, "en_stock"])

    records = []
    stock_disp = {}  # se reinicia por PT raíz

    def get_stk_info(comp):
        stk = df_stock[df_stock["cod"] == comp]
        en_stock    = float(stk["en_stock"].values[0])    if len(stk) else 0.0
        precio_prom = float(stk["precio_prom"].values[0]) if len(stk) else 0.0
        um          = stk["um"].values[0]                  if len(stk) else ""
        return en_stock, precio_prom, um

    def get_spec(comp):
        sp = df_spec[df_spec["cod"] == comp]
        tipo_mat  = sp["tipo_mat"].values[0]  if len(sp) else ""
        tipo_mat2 = sp["tipo_mat2"].values[0] if len(sp) else ""
        return str(tipo_mat or ""), str(tipo_mat2 or "")

    def recurse(cod_raiz, parent_cod, qty_neta, level):
        """qty_neta: cantidad real que hay que producir del padre (ya descontado stock)"""
        if parent_cod not in bom_dict:
            return
        qty_pt_base = qty_pt_map.get(parent_cod, 1.0) or 1.0
        for row in bom_dict[parent_cod]:
            comp      = row["cod_comp"]
            qty_unit  = row["qty_comp"] / qty_pt_base
            necesidad = qty_neta * qty_unit          # necesidad bruta del componente
            tipo      = classify(comp, cod_raiz)
            en_stock_orig, precio_prom, um = get_stk_info(comp)
            tipo_mat, tipo_mat2 = get_spec(comp)

            if not um:
                um = str(row["und_comp"] or "")

            if tipo == "Semiterminado":
                # ¿Cuánto stock disponible queda para este semi?
                disp = stock_disp.get(comp, 0.0)
                uso  = min(disp, necesidad)           # cuánto consumimos del stock
                neto = max(necesidad - disp, 0.0)     # cuánto hay que producir

                # actualizar stock disponible (consumido)
                stock_disp[comp] = max(disp - necesidad, 0.0)

                # registrar este semiterminado con su situación
                records.append({
                    "PT_Raiz":         cod_raiz,
                    "Nivel":           level,
                    "Tipo":            "Semiterminado",
                    "Accion":          "USAR STOCK" if neto == 0 else ("PRODUCIR PARCIAL" if uso > 0 else "PRODUCIR"),
                    "Cod_Componente":  comp,
                    "Desc_Componente": str(row["desc_comp"] or ""),
                    "Und":             um,
                    "Qty_Necesaria":   round(necesidad, 4),
                    "Stock_Disponible":round(disp, 4),
                    "Qty_De_Stock":    round(uso, 4),
                    "Qty_A_Producir":  round(neto, 4),
                    "Precio_Prom_S/":  round(precio_prom, 4),
                    "Valor_Nec_S/":    round(neto * precio_prom, 2),
                    "Tipo_Mat":        tipo_mat,
                    "Tipo_Mat2":       tipo_mat2,
                    "PT_Padre":        parent_cod,
                })

                # solo explotar si hay necesidad neta
                if neto > 0:
                    recurse(cod_raiz, comp, neto, level + 1)
            else:
                # material comprado: usar stock disponible (descontando lo ya asignado)
                disp_comp = stock_disp.get(comp, en_stock_orig)
                falta_comp = max(necesidad - disp_comp, 0)
                # consumir del stock disponible
                stock_disp[comp] = max(disp_comp - necesidad, 0.0)
                accion_comp = "COMPRAR" if falta_comp > 0 else "USAR STOCK"
                records.append({
                    "PT_Raiz":         cod_raiz,
                    "Nivel":           level,
                    "Tipo":            tipo,
                    "Accion":          accion_comp,
                    "Cod_Componente":  comp,
                    "Desc_Componente": str(row["desc_comp"] or ""),
                    "Und":             um,
                    "Qty_Necesaria":   round(necesidad, 4),
                    "Stock_Disponible":round(disp_comp, 4),
                    "Qty_A_Producir":  round(falta_comp, 4),
                    "Precio_Prom_S/":  round(precio_prom, 4),
                    "Valor_Nec_S/":    round(falta_comp * precio_prom, 2),
                    "Tipo_Mat":        tipo_mat,
                    "Tipo_Mat2":       tipo_mat2,
                    "PT_Padre":        parent_cod,
                })

    for pt, qty_pedida in pt_qty_map.items():
        if pt not in bom_dict:
            continue
        # Reiniciar stock por PT raíz para no contaminar entre distintos PTs
        stock_disp.clear()
        stock_disp.update(stock_inicial)
        recurse(pt, pt, float(qty_pedida), 1)

    return pd.DataFrame(records) if records else pd.DataFrame()


# ─── INICIO ──────────────────────────────────────────────────────────────────────
print("Cargando datos...")
try:
    DF_BOM, DF_STOCK, DF_SPEC, DF_TIEMPOS, DF_ORD = load_data(INPUT_PATH)
    DF_BOM["cod_pt"]   = DF_BOM["cod_pt"].astype(str).str.strip()
    DF_BOM["cod_comp"] = DF_BOM["cod_comp"].astype(str).str.strip()
    ALL_PT  = sorted(DF_BOM["cod_pt"].unique().tolist())
    PT_DESC = {c: str(DF_BOM[DF_BOM["cod_pt"]==c]["desc_pt"].iloc[0] or "")
               for c in ALL_PT}
    LOAD_ERROR = None
    print(f"OK — {len(ALL_PT)} PTs | muestra: {ALL_PT[:5]}")
except Exception as e:
    import traceback; traceback.print_exc()
    DF_BOM = DF_STOCK = DF_SPEC = DF_TIEMPOS = DF_ORD = None
    ALL_PT = []; PT_DESC = {}
    LOAD_ERROR = str(e)

# ─── APP ────────────────────────────────────────────────────────────────────────
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP],
                suppress_callback_exceptions=True)
app.title = "BOM Explorer — Layconsa"

SIDEBAR = {"position":"fixed","top":0,"left":0,"bottom":0,"width":"300px",
           "padding":"24px 18px","backgroundColor":"#1F4E79","overflowY":"auto","zIndex":100}
CONTENT = {"marginLeft":"320px","padding":"24px 28px","minHeight":"100vh","backgroundColor":"#F0F4F8"}

app.layout = html.Div([
    # SIDEBAR
    html.Div([
        html.Div([
            html.Span("BOM", style={"color":"#90CAF9","fontWeight":"700","fontSize":"22px"}),
            html.Span(" Explorer", style={"color":"white","fontWeight":"300","fontSize":"22px"}),
        ], style={"marginBottom":"4px"}),
        html.P("Layconsa · Plan Producción",
               style={"color":"#5B8DB8","fontSize":"11px","marginBottom":"28px"}),

        html.Label("Selección de PTs",
                   style={"color":"#90CAF9","fontSize":"11px","fontWeight":"700",
                          "letterSpacing":"0.08em","textTransform":"uppercase",
                          "marginBottom":"8px","display":"block"}),
        html.Button("📋  Ingresar códigos y cantidades", id="btn-open-modal",
                    style={"width":"100%","padding":"10px 14px","borderRadius":"8px",
                           "border":"1px solid #2E75B6","backgroundColor":"#163a5f",
                           "color":"white","fontSize":"13px","cursor":"pointer",
                           "textAlign":"left","marginBottom":"8px"}),
        html.Div(id="selected-chips", style={"marginBottom":"20px","minHeight":"40px"}),

        html.Hr(style={"borderColor":"#2E5F8A","margin":"16px 0"}),
        html.Label("Filtrar por tipo",
                   style={"color":"#90CAF9","fontSize":"11px","fontWeight":"700",
                          "letterSpacing":"0.08em","textTransform":"uppercase",
                          "marginBottom":"8px","display":"block"}),
        dcc.Checklist(id="filter-tipo",
            options=[{"label":" Semiterminado","value":"Semiterminado"},
                     {"label":" Comprado",     "value":"Comprado"},
                     {"label":" PT Útiles",    "value":"PT Útiles"},
                     {"label":" PT Cuadernos", "value":"PT Cuadernos"}],
            value=["Semiterminado","Comprado","PT Útiles","PT Cuadernos"],
            style={"color":"#CFD8DC","fontSize":"13px","lineHeight":"2.2"}),

        html.Hr(style={"borderColor":"#2E5F8A","margin":"16px 0"}),
        html.Button("⚡  Explotar BOM", id="btn-explotar",
                    style={"width":"100%","padding":"11px","borderRadius":"8px","border":"none",
                           "backgroundColor":"#1565C0","color":"white","fontWeight":"700",
                           "fontSize":"14px","cursor":"pointer","marginBottom":"10px"}),
        html.Button("⬇  Exportar Excel", id="btn-export",
                    style={"width":"100%","padding":"11px","borderRadius":"8px","border":"none",
                           "backgroundColor":"#1B5E20","color":"white","fontWeight":"700",
                           "fontSize":"14px","cursor":"pointer"}),
        dcc.Download(id="download-excel"),
        html.Div(id="sidebar-status",
                 style={"marginTop":"14px","fontSize":"11px","color":"#90CAF9","lineHeight":"1.7"}),
    ], style=SIDEBAR),

    # MODAL
    html.Div([
        html.Div([
            # Header
            html.Div([
                html.Span("📦  Selección múltiple — Códigos y Cantidades",
                          style={"fontWeight":"700","fontSize":"14px","color":"#1F4E79"}),
                html.Button("✕", id="btn-close-modal",
                            style={"background":"none","border":"none","fontSize":"18px",
                                   "cursor":"pointer","color":"#607D8B"}),
            ], style={"display":"flex","justifyContent":"space-between","alignItems":"center",
                      "padding":"16px 20px","borderBottom":"1px solid #E0E0E0","backgroundColor":"#F5F7FA"}),
            # Instrucciones
            html.Div([
                html.P("Pega desde Excel o escribe: código → TAB → cantidad (una fila por PT)",
                       style={"fontSize":"12px","color":"#546E7A","margin":"0 0 6px 0"}),
                html.Div([
                    html.Code("2110020095   1000", style={"display":"block","fontSize":"11px"}),
                    html.Code("2140370004   2080", style={"display":"block","fontSize":"11px"}),
                    html.P("Separadores aceptados: TAB, punto y coma, dos o más espacios",
                           style={"fontSize":"11px","color":"#90A4AE","margin":"6px 0 0"}),
                ], style={"backgroundColor":"#ECEFF1","borderRadius":"6px","padding":"8px 12px"}),
            ], style={"padding":"14px 20px","borderBottom":"1px solid #E0E0E0"}),
            # Cuerpo
            html.Div([
                html.Div([
                    html.Div([
                        html.Label("Lista de códigos",
                                   style={"fontSize":"11px","fontWeight":"700","color":"#546E7A",
                                          "textTransform":"uppercase","marginBottom":"6px","display":"block"}),
                        dcc.Textarea(id="modal-textarea",
                            placeholder="2110020095\t1000\n2140370004\t2080",
                            style={"width":"100%","height":"200px","fontSize":"12px",
                                   "fontFamily":"monospace","padding":"10px","borderRadius":"6px",
                                   "border":"1px solid #CFD8DC","resize":"vertical","lineHeight":"1.7"}),
                        html.Button("Procesar →", id="btn-parse-list",
                                    style={"marginTop":"10px","padding":"8px 18px","borderRadius":"6px",
                                           "border":"none","backgroundColor":"#1565C0","color":"white",
                                           "fontSize":"13px","fontWeight":"600","cursor":"pointer"}),
                        html.Div(id="debug-raw",
                                 style={"fontSize":"10px","color":"#FF7043","marginTop":"6px",
                                        "fontFamily":"monospace","wordBreak":"break-all"}),
                    ], style={"flex":"1","marginRight":"16px"}),
                    html.Div([
                        html.Label("Vista previa",
                                   style={"fontSize":"11px","fontWeight":"700","color":"#546E7A",
                                          "textTransform":"uppercase","marginBottom":"6px","display":"block"}),
                        html.Div(id="modal-preview",
                                 style={"height":"240px","overflowY":"auto","border":"1px solid #E0E0E0",
                                        "borderRadius":"6px","backgroundColor":"#FAFAFA"}),
                    ], style={"flex":"1"}),
                ], style={"display":"flex"}),
            ], style={"padding":"16px 20px","borderBottom":"1px solid #E0E0E0"}),
            # Footer
            html.Div([
                html.Div(id="modal-status", style={"fontSize":"12px","color":"#546E7A","flex":"1"}),
                html.Div([
                    html.Button("Cancelar", id="btn-modal-cancel",
                                style={"padding":"8px 18px","borderRadius":"6px",
                                       "border":"1px solid #CFD8DC","backgroundColor":"white",
                                       "color":"#546E7A","fontSize":"13px","cursor":"pointer",
                                       "marginRight":"8px"}),
                    html.Button("✓  Confirmar", id="btn-modal-confirm",
                                style={"padding":"8px 18px","borderRadius":"6px","border":"none",
                                       "backgroundColor":"#1565C0","color":"white","fontSize":"13px",
                                       "fontWeight":"700","cursor":"pointer"}),
                ]),
            ], style={"display":"flex","alignItems":"center","justifyContent":"space-between",
                      "padding":"14px 20px"}),
        ], style={"backgroundColor":"white","borderRadius":"12px","width":"680px","maxWidth":"95vw",
                  "boxShadow":"0 8px 40px rgba(0,0,0,0.22)","overflow":"hidden"}),
    ], id="modal-overlay",
       style={"display":"none","position":"fixed","top":0,"left":0,"right":0,"bottom":0,
              "backgroundColor":"rgba(0,0,0,0.55)","zIndex":1000,
              "alignItems":"center","justifyContent":"center"}),

    # CONTENT
    html.Div([
        html.Div(id="kpi-row", style={"display":"flex","gap":"14px","marginBottom":"22px","flexWrap":"wrap"}),
        dcc.Tabs(id="tabs", value="tab-niveles", children=[
            dcc.Tab(label="Explosión por Niveles", value="tab-niveles",
                    style={"fontWeight":"500","fontSize":"13px"},
                    selected_style={"fontWeight":"700","fontSize":"13px","borderTop":"3px solid #1565C0"}),
            dcc.Tab(label="Semiterminados (231x)", value="tab-semi",
                    style={"fontWeight":"500","fontSize":"13px"},
                    selected_style={"fontWeight":"700","fontSize":"13px","borderTop":"3px solid #2E7D32"}),
            dcc.Tab(label="Comprados", value="tab-comp",
                    style={"fontWeight":"500","fontSize":"13px"},
                    selected_style={"fontWeight":"700","fontSize":"13px","borderTop":"3px solid #E65100"}),
            dcc.Tab(label="Netos (Compras / Producir)", value="tab-netos",
                    style={"fontWeight":"500","fontSize":"13px"},
                    selected_style={"fontWeight":"700","fontSize":"13px","borderTop":"3px solid #00695C"}),

        ], style={"marginBottom":"16px"}),
        html.Div(id="tab-content"),
    ], style=CONTENT),

    dcc.Store(id="store-bom"),
    dcc.Store(id="store-pt-qty"),
    dcc.Store(id="store-modal-parsed"),
])

# ─── MODAL OPEN/CLOSE (clientside) ──────────────────────────────────────────────
app.clientside_callback(
    """
    function(o, c1, c2, c3) {
        const t = dash_clientside.callback_context.triggered[0].prop_id;
        const show = {display:'flex',position:'fixed',top:0,left:0,right:0,bottom:0,
                      backgroundColor:'rgba(0,0,0,0.55)',zIndex:1000,
                      alignItems:'center',justifyContent:'center'};
        const hide = {...show, display:'none'};
        return t.includes('btn-open-modal') ? show : hide;
    }
    """,
    Output("modal-overlay","style"),
    Input("btn-open-modal","n_clicks"),
    Input("btn-close-modal","n_clicks"),
    Input("btn-modal-cancel","n_clicks"),
    Input("btn-modal-confirm","n_clicks"),
    prevent_initial_call=True,
)

# ─── PARSEAR LISTA ───────────────────────────────────────────────────────────────
@app.callback(
    Output("modal-preview",      "children"),
    Output("modal-status",       "children"),
    Output("store-modal-parsed", "data"),
    Output("debug-raw",          "children"),
    Input("btn-parse-list","n_clicks"),
    State("modal-textarea","value"),
    prevent_initial_call=True,
)
def parse_list(n, text):
    if not text or not text.strip():
        return html.P("Sin datos", style={"padding":"12px","color":"#90A4AE","fontSize":"12px"}), "", {}, ""

    parsed = {}
    errors = []
    lines  = text.strip().splitlines()

    for i, raw_line in enumerate(lines, 1):
        # quitar caracteres invisibles excepto tab
        line = raw_line.replace("\r","").replace("\xa0"," ").strip()
        if not line:
            continue

        # --- separar código y cantidad ---
        # prioridad: tab > punto y coma > dos+ espacios > espacio simple (última palabra numérica)
        if "\t" in line:
            parts = [x.strip() for x in line.split("\t", 1)]
        elif ";" in line:
            parts = [x.strip() for x in line.split(";", 1)]
        else:
            # dos o más espacios
            m = re.split(r"  +", line, maxsplit=1)
            if len(m) == 2:
                parts = [x.strip() for x in m]
            else:
                # último token numérico
                toks = line.split()
                if len(toks) >= 2 and re.match(r"^[\d.,]+$", toks[-1]):
                    parts = [" ".join(toks[:-1]).strip(), toks[-1]]
                else:
                    parts = [line]

        cod = parts[0].strip()
        qty = 1.0

        if len(parts) == 2 and parts[1]:
            raw_q = parts[1].strip()
            # '14,500' → separador de miles → 14500
            # '1.234,56' → europeo → 1234.56
            # '1,234.56' → anglosajón → 1234.56
            if re.match(r"^\d{1,3}(\.\d{3})+(,\d*)?$", raw_q):
                # europeo: 1.234,56 o 1.234
                raw_q = raw_q.replace(".","").replace(",",".")
            elif re.match(r"^\d{1,3}(,\d{3})+(\.\d*)?$", raw_q):
                # anglosajón o miles sin decimal: 14,500 o 1,234.56
                raw_q = raw_q.replace(",","")
            else:
                # cualquier otra coma → decimal
                raw_q = raw_q.replace(",",".")
            try:
                qty = float(raw_q)
            except:
                errors.append(f"L{i}: cantidad inválida '{parts[1]}'")

        if not cod:
            continue

        # validar contra ALL_PT
        if cod in ALL_PT:
            parsed[cod] = qty
        elif cod.upper().startswith(("BESTEMP","TEMP")):
            parsed[cod] = qty
            errors.append(f"⚠ {cod}: temporal sin BOM")
        else:
            errors.append(f"No encontrado: '{cod}' (no existe en BOM)")

    # debug: mostrar repr de primera línea
    debug = f"repr línea 1: {repr(lines[0][:80])}" if lines else ""

    if not parsed:
        msg = f"0 válidos · {len(errors)} errores: {errors[:3]}"
        return (html.P(msg, style={"padding":"12px","color":"#EF5350","fontSize":"12px"}),
                msg, {}, debug)

    # Filas válidas
    valid_rows = []
    for cod, qty in parsed.items():
        desc = PT_DESC.get(cod, "—")
        valid_rows.append(html.Tr([
            html.Td(cod, style={"padding":"5px 10px","fontFamily":"monospace","fontSize":"12px",
                                "fontWeight":"600","color":"#1565C0"}),
            html.Td((desc[:38]+"…") if len(desc)>38 else desc,
                    style={"padding":"5px 10px","fontSize":"11px","color":"#546E7A"}),
            html.Td(f"{qty:,.0f}", style={"padding":"5px 10px","fontSize":"12px","fontWeight":"700",
                                          "textAlign":"right","color":"#1B5E20"}),
        ]))

    # Filas con error
    error_rows = []
    for err in errors:
        error_rows.append(html.Tr([
            html.Td("⚠", style={"padding":"5px 10px","fontSize":"13px","color":"#E65100"}),
            html.Td(err, style={"padding":"5px 10px","fontSize":"11px","color":"#BF360C",
                                "fontFamily":"monospace"}),
            html.Td("", style={"padding":"5px 10px"}),
        ]))

    thead_valid = html.Thead(html.Tr([
        html.Th("Código",      style={"padding":"6px 10px","fontSize":"11px","backgroundColor":"#E3F2FD","color":"#1565C0"}),
        html.Th("Descripción", style={"padding":"6px 10px","fontSize":"11px","backgroundColor":"#E3F2FD","color":"#1565C0"}),
        html.Th("Cantidad",    style={"padding":"6px 10px","fontSize":"11px","backgroundColor":"#E3F2FD","color":"#1565C0","textAlign":"right"}),
    ]))

    table = html.Div([
        # tabla válidos
        html.Table([thead_valid, html.Tbody(valid_rows)],
                   style={"width":"100%","borderCollapse":"collapse"}),
        # sección errores
        *([
            html.Div([
                html.Span(f"⚠  {len(errors)} código(s) no encontrado(s)",
                          style={"fontSize":"11px","fontWeight":"700","color":"#E65100",
                                 "display":"block","padding":"8px 10px 4px",
                                 "borderTop":"1px solid #FFE0B2","marginTop":"6px"}),
                html.Table([
                    html.Tbody(error_rows)
                ], style={"width":"100%","borderCollapse":"collapse",
                          "backgroundColor":"#FFF8F5"}),
            ]) if error_rows else html.Div()
        ]),
    ])

    status = f"✓ {len(parsed)} PT(s) válidos"
    if errors:
        status += f" · ⚠ {len(errors)} no encontrado(s)"

    return table, status, parsed, debug


# ─── CONFIRMAR ───────────────────────────────────────────────────────────────────
@app.callback(
    Output("store-pt-qty",   "data"),
    Output("selected-chips", "children"),
    Input("btn-modal-confirm","n_clicks"),
    State("store-modal-parsed","data"),
    prevent_initial_call=True,
)
def confirm(n, parsed):
    if not parsed:
        return {}, html.P("Sin selección", style={"color":"#546E7A","fontSize":"12px"})
    chips = []
    for cod, qty in parsed.items():
        desc  = PT_DESC.get(cod,"")
        short = (desc[:22]+"…") if len(desc)>22 else desc
        chips.append(html.Div([
            html.Div([
                html.Span(cod,   style={"fontFamily":"monospace","fontSize":"11px","fontWeight":"700","color":"#90CAF9","display":"block"}),
                html.Span(short, style={"fontSize":"10px","color":"#78909C","display":"block"}),
            ], style={"flex":"1"}),
            html.Span(f"{qty:,.0f}", style={"fontSize":"12px","fontWeight":"700","color":"#A5D6A7",
                                             "backgroundColor":"#1B3A2A","padding":"2px 7px",
                                             "borderRadius":"10px","marginLeft":"6px"}),
        ], style={"display":"flex","alignItems":"center","backgroundColor":"#163a5f",
                  "borderRadius":"8px","padding":"7px 10px","marginBottom":"5px",
                  "border":"1px solid #2E5F8A"}))
    return parsed, chips


# ─── EXPLOTAR ────────────────────────────────────────────────────────────────────
@app.callback(
    Output("store-bom",      "data"),
    Output("kpi-row",        "children"),
    Output("sidebar-status", "children"),
    Input("btn-explotar","n_clicks"),
    State("store-pt-qty","data"),
    prevent_initial_call=True,
)
def explotar(n, pt_qty):
    if LOAD_ERROR:
        return None, [], f"❌ {LOAD_ERROR}"
    if not pt_qty:
        return None, [], "⚠ Ingresa códigos primero"
    df = explode_bom(DF_BOM, pt_qty, DF_STOCK, DF_SPEC, PT_DESC)
    if df.empty:
        return None, [], "⚠ BOM vacía"

    n_uni  = df["Cod_Componente"].nunique()
    n_semi = df[df["Tipo"]=="Semiterminado"]["Cod_Componente"].nunique()
    n_comp = df[df["Tipo"]=="Comprado"]["Cod_Componente"].nunique()
    max_lv = int(df["Nivel"].max())
    # Solo materiales comprados que realmente hay que comprar (diferencia negativa)
    df_comprar = df[(df["Tipo"]=="Comprado") & (df["Diferencia"] < 0)]
    val_nec = df_comprar.drop_duplicates(["PT_Raiz","Cod_Componente"])["Valor_Nec_S/"].sum()

    def kpi(lbl, val, color):
        return html.Div([
            html.P(lbl, style={"fontSize":"10px","color":"#607D8B","margin":"0 0 4px","fontWeight":"700",
                               "textTransform":"uppercase","letterSpacing":"0.06em"}),
            html.P(str(val), style={"fontSize":"26px","fontWeight":"800","color":color,"margin":"0"}),
        ], style={"background":"white","borderRadius":"12px","padding":"14px 18px",
                  "boxShadow":"0 1px 3px rgba(0,0,0,0.08)","minWidth":"130px",
                  "borderTop":f"3px solid {color}"})

    kpis = [
        kpi("Componentes únicos", n_uni,  "#1565C0"),
        kpi("Semiterminados",     n_semi, "#2E7D32"),
        kpi("Comprados",          n_comp, "#E65100"),
        kpi("Niveles máx.",       max_lv, "#6A1B9A"),
        kpi("A comprar S/",       f"{val_nec:,.2f}", "#00695C"),
    ]
    return df.to_dict("records"), kpis, f"✓ {len(pt_qty)} PT(s) · {len(df)} líneas"


# ─── TABS ────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("tab-content","children"),
    Input("tabs","value"),
    Input("store-bom","data"),
    State("filter-tipo","value"),
    State("store-pt-qty","data"),
)
def render_tab(tab, data, tipos, pt_qty):
    if not data:
        return html.Div("Ingresa códigos y presiona Explotar BOM",
                        style={"color":"#90A4AE","fontSize":"14px","textAlign":"center","marginTop":"60px"})

    df = pd.DataFrame(data)
    if tipos:
        df = df[df["Tipo"].isin(tipos)]

    NUM_COLS = {"Qty_BOM_Base","Qty_Comp_BOM","Qty_Unit","Qty_Necesaria",
                "En_Stock","Diferencia","Precio_Prom_S/","Valor_Nec_S/",
                "Qty_Necesaria","Precio_Prom","Valor_Nec"}

    cond = [
        {"if":{"filter_query":'{Tipo} = "Semiterminado"'},"backgroundColor":"#E8F5E9","color":"#1B5E20"},
        {"if":{"filter_query":'{Tipo} = "Comprado"'},     "backgroundColor":"#FFF3E0","color":"#BF360C"},
        {"if":{"filter_query":'{Tipo} = "PT Útiles"'},    "backgroundColor":"#E3F2FD","color":"#0D47A1"},
        {"if":{"filter_query":'{Tipo} = "PT Cuadernos"'}, "backgroundColor":"#F3E5F5","color":"#4A148C"},
        {"if":{"column_id":"Nivel"}, "fontWeight":"700","textAlign":"center"},
        {"if":{"column_id":"Diferencia","filter_query":"{Diferencia} < 0"},
         "color":"#B71C1C","fontWeight":"700"},
        {"if":{"column_id":"Diferencia","filter_query":"{Diferencia} >= 0"},
         "color":"#1B5E20","fontWeight":"700"},
    ]

    COLS_NIV = ["PT_Raiz","Desc_PT_Raiz","Nivel","Tipo","Posicion","Cod_Componente","Desc_Componente",
                "Und","Qty_BOM_Base","Qty_Comp_BOM","Qty_Unit","Qty_Necesaria",
                "En_Stock","Diferencia","Precio_Prom_S/","Valor_Nec_S/","Tipo_Mat","Tipo_Mat2","CECO"]
    COLS_AGG = ["Cod_Componente","Desc_Componente","Und","Tipo_Mat","Tipo_Mat2",
                "Qty_Necesaria","En_Stock","Diferencia","Precio_Prom_S/","Valor_Nec_S/"]

    if tab == "tab-niveles":
        df_show = df[COLS_NIV].sort_values(["PT_Raiz","Nivel","Posicion"])
    elif tab == "tab-semi":
        g = df[df["Tipo"]=="Semiterminado"].groupby(
            ["Cod_Componente","Desc_Componente","Und","Tipo_Mat","Tipo_Mat2"],as_index=False)
        df_show = g.agg(Qty_Necesaria=("Qty_Necesaria","sum"),En_Stock=("En_Stock","first"),
                        Diferencia=("Diferencia","first"),
                        **{"Precio_Prom_S/":("Precio_Prom_S/","first")},
                        **{"Valor_Nec_S/":("Valor_Nec_S/","sum")}).sort_values("Cod_Componente")
        cond = [c for c in cond if "Tipo}" not in str(c)]
    elif tab == "tab-comp":
        g = df[df["Tipo"]=="Comprado"].groupby(
            ["Cod_Componente","Desc_Componente","Und","Tipo_Mat","Tipo_Mat2"],as_index=False)
        df_show = g.agg(Qty_Necesaria=("Qty_Necesaria","sum"),En_Stock=("En_Stock","first"),
                        Diferencia=("Diferencia","first"),
                        **{"Precio_Prom_S/":("Precio_Prom_S/","first")},
                        **{"Valor_Nec_S/":("Valor_Nec_S/","sum")}).sort_values("Cod_Componente")
        cond = [c for c in cond if "Tipo}" not in str(c)]
    else:  # tab-netos
        if not pt_qty or LOAD_ERROR:
            return html.P("Ingresa códigos y explota primero",
                          style={"color":"#90A4AE","fontSize":"13px","textAlign":"center","marginTop":"40px"})

        # Consolidar: semiterminados de calcular_netos + comprados de explosión directa
        df_n_raw = calcular_netos(DF_BOM, pt_qty, DF_STOCK, DF_SPEC)
        if df_n_raw.empty:
            return html.P("Sin resultados", style={"color":"#90A4AE","textAlign":"center","marginTop":"40px"})

        # Semiterminados consolidados
        df_semi_n = df_n_raw[df_n_raw["Tipo"]=="Semiterminado"].groupby(
            ["Cod_Componente","Desc_Componente","Und","Tipo","Accion","Tipo_Mat","Tipo_Mat2"],
            as_index=False
        ).agg(Qty_Necesaria=("Qty_Necesaria","sum"),
              Stock_Disponible=("Stock_Disponible","first"),
              Qty_A_Producir=("Qty_A_Producir","sum"),
              **{"Precio_Prom_S/":("Precio_Prom_S/","first")},
              **{"Valor_Nec_S/":("Valor_Nec_S/","sum")})

        # Comprados: desde df (explosión), necesidad total - stock real
        df_comp_n = df[df["Tipo"]=="Comprado"].groupby(
            ["Cod_Componente","Desc_Componente","Und","Tipo_Mat","Tipo_Mat2"],
            as_index=False
        ).agg(Qty_Necesaria=("Qty_Necesaria","sum"),
              Stock_Disponible=("En_Stock","first"),
              **{"Precio_Prom_S/":("Precio_Prom_S/","first")})
        df_comp_n["Qty_A_Producir"] = (df_comp_n["Qty_Necesaria"] - df_comp_n["Stock_Disponible"]).clip(lower=0).round(4)
        df_comp_n["Accion"]         = df_comp_n["Qty_A_Producir"].apply(lambda x: "COMPRAR" if x > 0 else "USAR STOCK")
        df_comp_n["Valor_Nec_S/"]   = (df_comp_n["Qty_A_Producir"] * df_comp_n["Precio_Prom_S/"]).round(2)
        df_comp_n["Tipo"]           = "Comprado"
        df_comp_n = df_comp_n[df_comp_n["Qty_A_Producir"] > 0]

        df_show_n = pd.concat([df_semi_n, df_comp_n], ignore_index=True).sort_values(["Accion","Cod_Componente"])

        NUM_N = {"Qty_Necesaria","Stock_Disponible","Qty_A_Producir","Precio_Prom_S/","Valor_Nec_S/"}
        ACCION_COLORS = {
            "USAR STOCK":      ("E8F5E9","1B5E20"),
            "PRODUCIR":        ("E3F2FD","0D47A1"),
            "PRODUCIR PARCIAL":("FFF9C4","F57F17"),
            "COMPRAR":         ("FFF3E0","BF360C"),
        }
        cond_n = []
        for ac, (bg, fg) in ACCION_COLORS.items():
            cond_n.append({"if":{"filter_query":f'{{Accion}} = "{ac}"'},
                           "backgroundColor":f"#{bg}","color":f"#{fg}"})
        cond_n += [
            {"if":{"column_id":"Qty_A_Producir","filter_query":"{Qty_A_Producir} > 0"},"fontWeight":"700"},
        ]
        col_defs_n = []
        for c in df_show_n.columns:
            if c in NUM_N:
                col_defs_n.append({"name":c,"id":c,"type":"numeric","format":{"specifier":",.4~f"}})
            else:
                col_defs_n.append({"name":c,"id":c})

        resumen = df_show_n.groupby("Accion", as_index=False).agg(
            Componentes=("Cod_Componente","nunique"),
            Valor=("Valor_Nec_S/","sum"))
        cards = html.Div([
            html.Div([
                html.P(r["Accion"], style={"fontSize":"10px","fontWeight":"700","margin":"0 0 4px",
                                           "color":f"#{ACCION_COLORS.get(r['Accion'],('607D8B','607D8B'))[1]}",
                                           "textTransform":"uppercase","letterSpacing":"0.06em"}),
                html.P(f"{int(r['Componentes'])} comp.",
                       style={"fontSize":"22px","fontWeight":"800","margin":"0",
                              "color":f"#{ACCION_COLORS.get(r['Accion'],('607D8B','607D8B'))[1]}"}),
                html.P(f"S/ {r['Valor']:,.2f}",
                       style={"fontSize":"11px","color":"#546E7A","margin":"4px 0 0"}),
            ], style={"background":"white","borderRadius":"10px","padding":"12px 16px",
                      "boxShadow":"0 1px 3px rgba(0,0,0,0.08)","minWidth":"150px",
                      "borderTop":f"3px solid #{ACCION_COLORS.get(r['Accion'],('607D8B','607D8B'))[1]}"})
            for _, r in resumen.iterrows()
        ], style={"display":"flex","gap":"12px","marginBottom":"16px","flexWrap":"wrap"})

        return html.Div([cards, dash_table.DataTable(
            data=df_show_n.to_dict("records"), columns=col_defs_n,
            page_size=60, sort_action="native", filter_action="native",
            style_table={"overflowX":"auto","borderRadius":"12px","overflow":"hidden",
                         "boxShadow":"0 1px 4px rgba(0,0,0,0.08)"},
            style_header={"backgroundColor":"#004D40","color":"white","fontWeight":"700",
                          "fontSize":"11px","padding":"10px 12px","border":"none"},
            style_cell={"fontSize":"12px","padding":"7px 11px","border":"1px solid #ECEFF1",
                        "fontFamily":"Arial, sans-serif","whiteSpace":"normal","maxWidth":"280px"},
            style_cell_conditional=[{"if":{"column_id":c},"textAlign":"right","fontFamily":"monospace","fontSize":"11px"}
                                     for c in NUM_N],
            style_data_conditional=cond_n,
            style_data={"backgroundColor":"white"},
        )])

    col_defs = []
    for c in df_show.columns:
        if c in NUM_COLS:
            col_defs.append({"name":c,"id":c,"type":"numeric","format":{"specifier":",.4~f"}})
        else:
            col_defs.append({"name":c,"id":c})

    return dash_table.DataTable(
        data=df_show.to_dict("records"), columns=col_defs,
        page_size=60, sort_action="native", filter_action="native",
        style_table={"overflowX":"auto","borderRadius":"12px","overflow":"hidden",
                     "boxShadow":"0 1px 4px rgba(0,0,0,0.08)"},
        style_header={"backgroundColor":"#1F4E79","color":"white","fontWeight":"700",
                      "fontSize":"11px","padding":"10px 12px","border":"none"},
        style_cell={"fontSize":"12px","padding":"7px 11px","border":"1px solid #ECEFF1",
                    "fontFamily":"Arial, sans-serif","whiteSpace":"normal","maxWidth":"280px"},
        style_cell_conditional=[{"if":{"column_id":c},"textAlign":"right","fontFamily":"monospace","fontSize":"11px"}
                                 for c in NUM_COLS],
        style_data_conditional=cond,
        style_data={"backgroundColor":"white"},
    )



# ─── EXPORT ──────────────────────────────────────────────────────────────────────
@app.callback(
    Output("download-excel","data"),
    Input("btn-export","n_clicks"),
    State("store-bom","data"),
    State("store-pt-qty","data"),
    prevent_initial_call=True,
)
def export_excel(n, data, pt_qty):
    if not data:
        return dash.no_update
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(data)
    df_semi = df[df["Tipo"]=="Semiterminado"].groupby(
        ["Cod_Componente","Desc_Componente","Und","Tipo_Mat","Tipo_Mat2"],as_index=False
    ).agg(Qty_Necesaria=("Qty_Necesaria","sum"),En_Stock=("En_Stock","first"),
          Diferencia=("Diferencia","first"),
          **{"Precio_Prom_S/":("Precio_Prom_S/","first")},
          **{"Valor_Nec_S/":("Valor_Nec_S/","sum")}).sort_values("Cod_Componente")
    df_comp = df[df["Tipo"]=="Comprado"].groupby(
        ["Cod_Componente","Desc_Componente","Und","Tipo_Mat","Tipo_Mat2"],as_index=False
    ).agg(Qty_Necesaria=("Qty_Necesaria","sum"),En_Stock=("En_Stock","first"),
          Diferencia=("Diferencia","first"),
          **{"Precio_Prom_S/":("Precio_Prom_S/","first")},
          **{"Valor_Nec_S/":("Valor_Nec_S/","sum")}).sort_values("Cod_Componente")

    # Hoja 4 — Netos consolidado
    # Lógica correcta para comprados compartidos entre PTs:
    # 1. Sumar Qty_Necesaria total por componente (de todos los PTs)
    # 2. Restar el stock real UNA SOLA VEZ → Qty_A_Producir = max(total_necesario - stock_real, 0)
    df_netos = pd.DataFrame()
    if pt_qty and not df.empty:
        # Usar la explosión completa (df) que tiene todas las necesidades correctas
        # Agrupar comprados y semiterminados por separado

        # Semiterminados: viene de calcular_netos (ya tiene la lógica correcta de stock)
        df_netos_raw = calcular_netos(DF_BOM, pt_qty, DF_STOCK, DF_SPEC)

        if not df_netos_raw.empty:
            # ── Semiterminados: agregar desde netos_raw ──────────────────────
            df_semi_netos = df_netos_raw[df_netos_raw["Tipo"]=="Semiterminado"].groupby(
                ["Cod_Componente","Desc_Componente","Und","Tipo","Accion","Tipo_Mat","Tipo_Mat2"],
                as_index=False
            ).agg(
                Qty_Necesaria  =("Qty_Necesaria",   "sum"),
                Stock_Disponible=("Stock_Disponible","first"),
                Qty_A_Producir =("Qty_A_Producir",  "sum"),
                **{"Precio_Prom_S/": ("Precio_Prom_S/","first")},
                **{"Valor_Nec_S/":   ("Valor_Nec_S/",  "sum")},
            )

            # ── Comprados: consolidar desde explosión y restar stock real ────
            df_comp_exp = df[df["Tipo"]=="Comprado"].groupby(
                ["Cod_Componente","Desc_Componente","Und","Tipo_Mat","Tipo_Mat2"],
                as_index=False
            ).agg(
                Qty_Necesaria=("Qty_Necesaria","sum"),
                En_Stock     =("En_Stock","first"),
                **{"Precio_Prom_S/": ("Precio_Prom_S/","first")},
            )
            df_comp_exp["Stock_Disponible"] = df_comp_exp["En_Stock"]
            df_comp_exp["Qty_A_Producir"]   = (df_comp_exp["Qty_Necesaria"] - df_comp_exp["En_Stock"]).clip(lower=0).round(4)
            df_comp_exp["Accion"]           = df_comp_exp["Qty_A_Producir"].apply(lambda x: "COMPRAR" if x > 0 else "USAR STOCK")
            df_comp_exp["Valor_Nec_S/"]     = (df_comp_exp["Qty_A_Producir"] * df_comp_exp["Precio_Prom_S/"]).round(2)
            df_comp_exp["Tipo"]             = "Comprado"
            df_comp_exp = df_comp_exp.drop(columns=["En_Stock"])
            # Solo los que hay que comprar
            df_comp_netos = df_comp_exp[df_comp_exp["Qty_A_Producir"] > 0].sort_values("Cod_Componente")

            # Concatenar
            df_netos = pd.concat([df_semi_netos, df_comp_netos], ignore_index=True)                         .sort_values(["Accion","Cod_Componente"])

    # ── Hoja 4 Netos: agregar columnas de tiempo ──────────────────────────────────
    if not df_netos.empty and DF_TIEMPOS is not None and not DF_TIEMPOS.empty:
        def safe_div(a, b): return round(a / b, 4) if b and b > 0 else 0
        # deduplicar tiempos: 1 fila por código (primera ocurrencia — nivel proceso principal)
        df_t = DF_TIEMPOS[["cod_semi","cant_base","t_maq","maquina","proceso"]]                   .drop_duplicates("cod_semi", keep="first")
        df_netos = df_netos.merge(df_t, left_on="Cod_Componente", right_on="cod_semi", how="left")                           .drop(columns="cod_semi")
        df_netos["Horas_Maq_Nec"]  = df_netos.apply(
            lambda r: safe_div(r["Qty_A_Producir"] * r["t_maq"], r["cant_base"])
            if pd.notna(r.get("cant_base")) and r.get("cant_base",0) > 0 else 0, axis=1)
        df_netos["Dias_Produccion"] = df_netos["Horas_Maq_Nec"].apply(
            lambda h: round(h / 24, 2) if h > 0 else 0)
        # solo columnas requeridas
        cols_final = ["Cod_Componente","Desc_Componente","Und","Tipo","Accion","Tipo_Mat","Tipo_Mat2",
                      "Qty_Necesaria","Stock_Disponible","Qty_A_Producir","Precio_Prom_S/","Valor_Nec_S/",
                      "maquina","proceso","cant_base","Horas_Maq_Nec","Dias_Produccion"]
        cols_final = [c for c in cols_final if c in df_netos.columns]
        df_netos = df_netos[cols_final].rename(columns={
            "maquina":"Maquina","proceso":"Proceso","cant_base":"Cant_Base"})

    # ── Hoja 4 Netos: enriquecer comprados con MOQ/LT/tipo_compra ─────────────────
    if not df_netos.empty and DF_ORD is not None and not DF_ORD.empty:
        # Asegurar que columnas existen antes del merge
        for col in ["Cant_Base","Proceso","Dias_Produccion","Maquina","Horas_Maq_Nec"]:
            if col not in df_netos.columns:
                df_netos[col] = None

        df_netos = df_netos.merge(DF_ORD, left_on="Cod_Componente", right_on="cod", how="left").drop(columns="cod")

        # Para comprados: rellenar con datos de Ordenes
        mask_comp = df_netos["Tipo"] != "Semiterminado"
        df_netos.loc[mask_comp, "Cant_Base"]       = df_netos.loc[mask_comp, "moq"]
        df_netos.loc[mask_comp, "Proceso"]         = df_netos.loc[mask_comp, "tipo_compra"]
        df_netos.loc[mask_comp, "Dias_Produccion"] = df_netos.loc[mask_comp, "lt_dias"]
        df_netos = df_netos.drop(columns=["moq","tipo_compra","lt_dias"], errors="ignore")

        # Reordenar columnas para que quede limpio
        cols_ord = ["Cod_Componente","Desc_Componente","Und","Tipo","Accion","Tipo_Mat","Tipo_Mat2",
                    "Qty_Necesaria","Stock_Disponible","Qty_A_Producir","Precio_Prom_S/","Valor_Nec_S/",
                    "Maquina","Proceso","Cant_Base","Horas_Maq_Nec","Dias_Produccion"]
        cols_ord = [c for c in cols_ord if c in df_netos.columns]
        df_netos = df_netos[cols_ord]

    # ── Hoja 5 Gantt ─────────────────────────────────────────────────────────────
    # Construir tabla Gantt: solo semiterminados 231 que faltan, con sus comprados debajo
    # Scheduling: nivel más profundo = Día 1, cada nivel espera al más lento del anterior

    def build_gantt(df_exp_full, df_netos_raw, df_tiempos, df_ord_data, pt_desc_map):
        """
        Lógica correcta:
        - Usar calcular_netos para saber QUÉ semis realmente hay que producir
          (descarta semis con stock suficiente)
        - Scheduling por niveles: nivel más profundo = Día 1
        - PT aparece al final con su puesto de Tiempos
        - Materiales comprados debajo de cada semi (todos, para ver stock y LT)
        """
        # Semiterminados que realmente hay que producir (de calcular_netos)
        semis_producir = set()
        if df_netos_raw is not None and not df_netos_raw.empty:
            mask = (df_netos_raw["Tipo"]=="Semiterminado") & (df_netos_raw["Qty_A_Producir"] > 0)
            semis_producir = set(df_netos_raw[mask]["Cod_Componente"].unique())

        # Diccionarios de duración y máquina
        dur_map = {}   # cod → dias
        maq_map = {}   # cod → maquina

        if df_tiempos is not None and not df_tiempos.empty:
            df_t = df_tiempos.drop_duplicates("cod_semi", keep="first")
            # Calcular duración para cada semi que hay que producir
            qty_map = {}
            if df_netos_raw is not None and not df_netos_raw.empty:
                for _, r in df_netos_raw[df_netos_raw["Tipo"]=="Semiterminado"].iterrows():
                    cod = r["Cod_Componente"]
                    qty_map[cod] = qty_map.get(cod, 0) + float(r.get("Qty_A_Producir", 0))
            # También PTs desde pt_qty_map
            for _, r in df_exp_full[df_exp_full["Tipo"].isin(["PT Útiles","PT Cuadernos"])].drop_duplicates("PT_Raiz").iterrows():
                pass  # PTs se manejan abajo

            for cod in set(df_t["cod_semi"]):
                row_t = df_t[df_t["cod_semi"]==cod].iloc[0]
                cb = float(row_t.get("cant_base") or 1)
                tm = float(row_t.get("t_maq") or 0)
                maq_map[cod] = str(row_t.get("maquina","") or "Sin máquina")
                qty = qty_map.get(cod, 0)
                if qty > 0 and cb > 0 and tm > 0:
                    dias = round((qty * tm / cb) / 24, 2)
                    dur_map[cod] = max(dias, 0.5)
                elif tm > 0 and cb > 0:
                    dur_map[cod] = 0  # tiene datos pero qty=0

        if df_ord_data is not None and not df_ord_data.empty:
            for _, r in df_ord_data.iterrows():
                lt = r["lt_dias"]
                if pd.notna(lt):
                    try: dur_map[str(r["cod"])] = float(lt)
                    except: pass

        gantt_rows = []

        for pt_raiz in df_exp_full["PT_Raiz"].unique():
            desc_pt = pt_desc_map.get(pt_raiz, "")

            # ── Encabezado del PT ──────────────────────────────────────────────
            gantt_rows.append({
                "Fila_Tipo":"PT_HDR","PT_Raiz":pt_raiz,"Desc_PT":desc_pt,
                "Nivel_BOM":"","Puesto":"","Cod_Componente":pt_raiz,
                "Desc_Componente":desc_pt,"Und":"","Qty_A_Producir":"",
                "Stock_Semi":"","Duracion_dias":"","Dia_Inicio":"","Dia_Fin":"",
                "LT_dias":"","Stock_Mat":"","Falta_Mat":"",
            })

            # Semis de este PT que hay que producir, filtrados por calcular_netos
            # Solo semis que calcular_netos marcó como PRODUCIR o PRODUCIR PARCIAL
            # Esto excluye semis cuyos padres tienen stock (como 2310180048 cuando la Tinta tiene stock)
            df_pt_exp = df_exp_full[
                (df_exp_full["PT_Raiz"]==pt_raiz) &
                (df_exp_full["Tipo"]=="Semiterminado") &
                (df_exp_full["Cod_Componente"].isin(semis_producir)) &
                (df_exp_full["Diferencia"] < 0)
            ].copy()

            # Niveles presentes (de más profundo a más superficial)
            niveles = sorted(df_pt_exp["Nivel"].unique(), reverse=True) if not df_pt_exp.empty else []

            # ── Calcular scheduling por nivel ──────────────────────────────────
            # Dentro de cada nivel: paralelo → el nivel termina cuando termina el más lento
            # Comprados de cualquier nivel también desde Día 1 (paralelo a todo)
            # El nivel N espera al max(fin nivel N+1, fin comprados críticos de nivel N)
            nivel_fin = {}
            dia_acum = 1
            for nivel in niveles:
                df_nv = df_pt_exp[df_pt_exp["Nivel"]==nivel]
                max_dur = max((dur_map.get(r["Cod_Componente"], 0) for _, r in df_nv.iterrows()), default=0)
                # También considerar LT de comprados hijos de este nivel
                for _, r in df_nv.iterrows():
                    hijos_comp = df_exp_full[
                        (df_exp_full["PT_Padre"]==r["Cod_Componente"]) &
                        (df_exp_full["Tipo"]=="Comprado") &
                        (df_exp_full["Diferencia"] < 0)
                    ]["Cod_Componente"].unique()
                    for hc in hijos_comp:
                        lt_h = dur_map.get(hc, 0)
                        if lt_h > max_dur:
                            max_dur = lt_h
                nivel_fin[nivel] = dia_acum + max_dur
                dia_acum = nivel_fin[nivel] + 1

            # ── Filas semiterminados (de nivel más profundo a superficial) ─────
            semis_vistos = set()
            for nivel in sorted(niveles, reverse=True) if niveles else []:  # nivel 3→2→1
                df_nv = df_pt_exp[df_pt_exp["Nivel"]==nivel]
                for _, r in df_nv.iterrows():
                    cod_semi = r["Cod_Componente"]
                    if cod_semi in semis_vistos:
                        continue
                    semis_vistos.add(cod_semi)

                    dur = dur_map.get(cod_semi, 0)
                    maq = maq_map.get(cod_semi, "Sin máquina")
                    # Inicio: cuando terminan los niveles más profundos
                    # LT máximo de comprados hijos de ESTE semi que faltan
                    lt_hijos = [
                        dur_map.get(hc, 0)
                        for hc in df_exp_full[
                            (df_exp_full["PT_Padre"]==cod_semi) &
                            (df_exp_full["Tipo"]=="Comprado") &
                            (df_exp_full["Diferencia"] < 0)
                        ]["Cod_Componente"].unique()
                    ]
                    max_lt_hijos = max(lt_hijos) if lt_hijos else 0

                    nivs_prof = [n for n in niveles if n > nivel]
                    if nivs_prof:
                        dia_ini_semi = max(nivel_fin[max(nivs_prof)] + 1, max_lt_hijos + 1)
                    else:
                        dia_ini_semi = max_lt_hijos + 1 if max_lt_hijos > 0 else 1

                    # Obtener qty_a_producir desde netos
                    qty_prod = 0
                    if df_netos_raw is not None and not df_netos_raw.empty:
                        nr = df_netos_raw[df_netos_raw["Cod_Componente"]==cod_semi]
                        if len(nr): qty_prod = float(nr.iloc[0].get("Qty_A_Producir", 0))
                    stock_s = float(r["En_Stock"])
                    dia_fin = round(dia_ini_semi + dur, 1)

                    gantt_rows.append({
                        "Fila_Tipo":"SEMI","PT_Raiz":pt_raiz,"Desc_PT":desc_pt,
                        "Nivel_BOM":nivel,"Puesto":maq,
                        "Cod_Componente":cod_semi,
                        "Desc_Componente":str(r["Desc_Componente"]),
                        "Und":str(r["Und"]),
                        "Qty_A_Producir":round(qty_prod, 2),
                        "Stock_Semi":round(stock_s, 2),
                        "Duracion_dias":round(dur, 2),
                        "Dia_Inicio":dia_ini_semi,
                        "Dia_Fin":dia_fin,
                        "LT_dias":"","Stock_Mat":"","Falta_Mat":"",
                    })

                    # Comprados hijos (todos, para ver stock y LT)
                    hijos_comp_df = df_exp_full[
                        (df_exp_full["PT_Padre"]==cod_semi) &
                        (df_exp_full["Tipo"]=="Comprado")
                    ].drop_duplicates("Cod_Componente")
                    for _, h in hijos_comp_df.iterrows():
                        lt_v = dur_map.get(h["Cod_Componente"], None)
                        falta = float(h["Diferencia"]) if float(h["Diferencia"]) < 0 else ""
                        gantt_rows.append({
                            "Fila_Tipo":"MAT","PT_Raiz":pt_raiz,"Desc_PT":desc_pt,
                            "Nivel_BOM":nivel,"Puesto":"",
                            "Cod_Componente":f"  └ {h['Cod_Componente']}",
                            "Desc_Componente":f"    {h['Desc_Componente']}",
                            "Und":str(h["Und"]),
                            "Qty_A_Producir":"","Stock_Semi":"",
                            "Duracion_dias":"","Dia_Inicio":"","Dia_Fin":"",
                            "LT_dias":lt_v if lt_v else "",
                            "Stock_Mat":round(float(h["En_Stock"]), 2),
                            "Falta_Mat":round(float(falta), 2) if falta != "" else "",
                        })

            # ── Comprados directos del PT (nivel 1, hijos del PT no de ningún semi) ──
            comp_directos_pt = df_exp_full[
                (df_exp_full["PT_Padre"]==pt_raiz) &
                (df_exp_full["Tipo"]=="Comprado")
            ].drop_duplicates("Cod_Componente")

            for _, h in comp_directos_pt.iterrows():
                lt_v  = dur_map.get(h["Cod_Componente"], None)
                falta = float(h["Diferencia"]) if float(h["Diferencia"]) < 0 else ""
                gantt_rows.append({
                    "Fila_Tipo":"MAT_PT","PT_Raiz":pt_raiz,"Desc_PT":desc_pt,
                    "Nivel_BOM":1,"Puesto":"",
                    "Cod_Componente":f"  └ {h['Cod_Componente']}",
                    "Desc_Componente":f"    {h['Desc_Componente']}",
                    "Und":str(h["Und"]),
                    "Qty_A_Producir":"","Stock_Semi":"",
                    "Duracion_dias":"","Dia_Inicio":"","Dia_Fin":"",
                    "LT_dias":lt_v if lt_v else "",
                    "Stock_Mat":round(float(h["En_Stock"]), 2),
                    "Falta_Mat":round(float(falta), 2) if falta != "" else "",
                })

            # ── PT al final ────────────────────────────────────────────────────
            # Puesto del PT desde Tiempos (buscar por cod_semi == pt_raiz)
            maq_pt = "Ensamblado"
            if df_tiempos is not None and not df_tiempos.empty:
                pt_t = df_tiempos[df_tiempos["cod_semi"]==pt_raiz]
                if len(pt_t): maq_pt = str(pt_t.iloc[0].get("maquina","") or "Ensamblado")

            # PT empieza cuando termina el nivel 1 (más superficial = min nivel)
            # Y también cuando llegan sus comprados directos que faltan
            # Si hay niveles de semis → esperar al más superficial
            # Si no hay semis que producir (todo en stock) → arrancar directo
            if niveles:
                nivel_1 = min(niveles)
                dia_base_pt = nivel_fin.get(nivel_1, 1) + 1
            else:
                dia_base_pt = 1

            lt_comp_directos = [
                dur_map.get(h["Cod_Componente"], 0)
                for _, h in comp_directos_pt[comp_directos_pt["Diferencia"] < 0].iterrows()
            ]
            max_lt_directos = max(lt_comp_directos) if lt_comp_directos else 0
            dia_ini_pt = max(dia_base_pt, max_lt_directos + 1) if max_lt_directos > 0 else dia_base_pt
            dur_pt = dur_map.get(pt_raiz, 0)
            # Si no hay duración del PT en tiempos, usar 1 día por defecto
            if dur_pt == 0:
                if df_tiempos is not None and not df_tiempos.empty:
                    pt_t = df_tiempos[df_tiempos["cod_semi"]==pt_raiz]
                    if len(pt_t):
                        cb = float(pt_t.iloc[0].get("cant_base") or 1)
                        tm = float(pt_t.iloc[0].get("t_maq") or 0)
                        # qty del PT = lo que se pidió producir
                        qty_pt_v = sum(float(v) for v in [pt_qty.get(pt_raiz, 1)] if v)
                        dur_pt = round((qty_pt_v * tm / cb) / 24, 2) if cb > 0 and tm > 0 else 1.0
                dur_pt = max(dur_pt, 1.0)

            gantt_rows.append({
                "Fila_Tipo":"PT","PT_Raiz":pt_raiz,"Desc_PT":desc_pt,
                "Nivel_BOM":0,"Puesto":maq_pt,
                "Cod_Componente":pt_raiz,
                "Desc_Componente":f"► {desc_pt}",
                "Und":"","Qty_A_Producir":pt_qty.get(pt_raiz,""),
                "Stock_Semi":"","Duracion_dias":round(dur_pt, 2),
                "Dia_Inicio":dia_ini_pt,"Dia_Fin":round(dia_ini_pt+dur_pt, 1),
                "LT_dias":"","Stock_Mat":"","Falta_Mat":"",
            })

        return pd.DataFrame(gantt_rows)

    df_gantt = pd.DataFrame()
    if pt_qty:
        df_exp_full  = explode_bom(DF_BOM, pt_qty, DF_STOCK, DF_SPEC, PT_DESC)
        df_netos_raw = calcular_netos(DF_BOM, pt_qty, DF_STOCK, DF_SPEC) if not df_exp_full.empty else pd.DataFrame()
        if not df_exp_full.empty:
            df_gantt = build_gantt(df_exp_full, df_netos_raw, DF_TIEMPOS, DF_ORD, PT_DESC)

    # ── Hoja 0 Resumen ───────────────────────────────────────────────────────────
    resumen_rows = []
    for pt, qty in (pt_qty or {}).items():
        desc = PT_DESC.get(pt, "")
        # ¿Se puede hacer? → verificar si hay materiales faltantes sin LT o con LT crítico
        faltantes = 0
        if not df_gantt.empty:
            faltantes = len(df_gantt[
                (df_gantt["PT_Raiz"]==pt) &
                (df_gantt["Fila_Tipo"]=="MAT") &
                (df_gantt["Falta_Mat"] != "")
            ])
        se_puede = "Sí" if faltantes == 0 else ("Parcial" if df_gantt[
            (df_gantt["PT_Raiz"]==pt) & (df_gantt["Fila_Tipo"]=="PT")
        ]["Qty_A_Producir"].sum() > 0 else "No")

        # Para cuando → Dia_Fin del PT
        para_cuando = ""
        if not df_gantt.empty:
            pt_row = df_gantt[(df_gantt["PT_Raiz"]==pt) & (df_gantt["Fila_Tipo"]=="PT")]
            if len(pt_row):
                para_cuando = pt_row.iloc[0]["Dia_Fin"]

        stk_pt = DF_STOCK[DF_STOCK["cod"]==pt]["en_stock"].sum() if DF_STOCK is not None else 0
        resumen_rows.append({
    "Codigo":        pt,
    "Descripcion":   desc,
    "Cantidad":      qty,
    "Stock_PT":      round(stk_pt, 2),
    "Se_puede_hacer":se_puede,
    "Para_cuando_dia":para_cuando,
    "Cuanto":        "",
    "Limitante":     "",
})
    df_resumen = pd.DataFrame(resumen_rows)

    # ── Hoja Resumen Gerencial (una fila por PT con barra Gantt) ─────────────────
    df_resumen_gantt = pd.DataFrame()
    if not df_gantt.empty:
        pt_rows = df_gantt[df_gantt["Fila_Tipo"]=="PT"].copy()
        if not pt_rows.empty:
            df_resumen_gantt = pt_rows[[
                "PT_Raiz","Desc_PT","Qty_A_Producir","Duracion_dias","Dia_Inicio","Dia_Fin"
            ]].rename(columns={
                "PT_Raiz":       "Codigo",
                "Desc_PT":       "Descripcion",
                "Qty_A_Producir":"Cantidad",
                "Duracion_dias": "Duracion_dias",
                "Dia_Inicio":    "Dia_Inicio",
                "Dia_Fin":       "Dia_Fin",
            }).reset_index(drop=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        if not df_resumen.empty:
            df_resumen.to_excel(w, sheet_name="0_Resumen",          index=False)
        if not df_resumen_gantt.empty:
            df_resumen_gantt.to_excel(w, sheet_name="0_Plan_Produccion", index=False)
        df.sort_values(["PT_Raiz","Nivel"]).to_excel(w, sheet_name="1_Explosion_Niveles", index=False)
        df_semi.to_excel(w, sheet_name="2_Semiterminados",  index=False)
        df_comp.to_excel(w, sheet_name="3_Comprados",       index=False)
        if not df_netos.empty:
            df_netos.to_excel(w, sheet_name="4_Netos",      index=False)
        if not df_gantt.empty:
            df_gantt.to_excel(w, sheet_name="5_Gantt",      index=False)
    buf.seek(0)
    wb2 = load_workbook(buf)

    TIPO_FILL = {"Semiterminado":"C8E6C9","Comprado":"FFE0B2",
                 "PT Útiles":"BBDEFB","PT Cuadernos":"E1BEE7"}
    HDR = {"0_Resumen":"880E4F","0_Plan_Produccion":"004D40","1_Explosion_Niveles":"1F4E79",
           "2_Semiterminados":"1B5E20","3_Comprados":"7B3F00","4_Netos":"004D40","5_Gantt":"1A237E"}
    ACCION_FILLS = {"USAR STOCK":"C8E6C9","PRODUCIR":"BBDEFB","PRODUCIR PARCIAL":"FFF9C4","COMPRAR":"FFE0B2"}
    NUM_NAMES = {"Qty_BOM_Base","Qty_Comp_BOM","Qty_Unit","Qty_Necesaria","Qty_A_Producir",
                 "En_Stock","Diferencia","Precio_Prom_S/","Valor_Nec_S/","Precio_Prom","Valor_Nec",
                 "Cant_Base","Horas_Maq_Nec","Dias_Produccion","Duracion_dias","Dia_Inicio","Dia_Fin",
                 "Duracion_total","N_Actividades","MOQ","LT_dias"}
    thin = Side(style="thin", color="D0D0D0")
    brd  = Border(left=thin,right=thin,top=thin,bottom=thin)

    for sn in wb2.sheetnames:
        ws = wb2[sn]
        hdrs = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        num_idx = {i+1 for i,h in enumerate(hdrs) if h and str(h) in NUM_NAMES}
        tipo_idx = (hdrs.index("Tipo")+1) if "Tipo" in hdrs else None
        dif_idx  = (hdrs.index("Diferencia")+1) if "Diferencia" in hdrs else None

        for c in range(1, ws.max_column+1):
            cell = ws.cell(1,c)
            cell.font      = Font(bold=True,color="FFFFFF",size=10,name="Arial")
            cell.fill      = PatternFill("solid",fgColor=HDR.get(sn,"1F4E79"))
            cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border    = brd
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        for r in range(2, ws.max_row+1):
            tipo_val  = ws.cell(r, tipo_idx).value if tipo_idx else ""
            # Para hoja netos usar color por Accion
            if sn == "4_Netos" and "Accion" in hdrs:
                accion_idx = hdrs.index("Accion") + 1
                accion_val = ws.cell(r, accion_idx).value
                fill_hex = ACCION_FILLS.get(str(accion_val), "FFFFFF")
            else:
                fill_hex  = TIPO_FILL.get(str(tipo_val), "FFFFFF")
            for c in range(1, ws.max_column+1):
                cell = ws.cell(r,c)
                cell.border = brd
                cell.font   = Font(size=9,name="Arial")
                cell.fill   = PatternFill("solid",fgColor=fill_hex)
                if c in num_idx:
                    cell.number_format = "#,##0.000"
                    cell.alignment = Alignment(horizontal="right")
                    try: cell.value = float(cell.value)
                    except: pass
                if dif_idx and c == dif_idx:
                    try:
                        v = float(cell.value)
                        cell.font = Font(size=9,name="Arial",bold=True,
                                        color="B71C1C" if v < 0 else "1B5E20")
                    except: pass

        for c in range(1, ws.max_column+1):
            col_l = get_column_letter(c)
            mx = max((len(str(ws.cell(r,c).value or "")) for r in range(1, min(ws.max_row+1,50))), default=8)
            ws.column_dimensions[col_l].width = min(max(mx+2, 10), 55)
        ws.auto_filter.ref = ws.dimensions



    # ── Formato hoja Plan_Produccion (resumen gerencial con barras) ─────────────
    if "0_Plan_Produccion" in wb2.sheetnames:
        ws_p = wb2["0_Plan_Produccion"]
        hdrs_p = [ws_p.cell(1,c).value for c in range(1, ws_p.max_column+1)]

        # Encabezado
        for c in range(1, ws_p.max_column+1):
            cell = ws_p.cell(1,c)
            cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Arial")
            cell.fill      = PatternFill("solid", fgColor="004D40")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = brd
        ws_p.row_dimensions[1].height = 32

        DATA_COLS_P = len(hdrs_p)
        MAX_DIAS_P  = 120
        BASE_COL_P  = DATA_COLS_P + 2

        # Encabezados de días
        ws_p.cell(1, DATA_COLS_P+1).value = ""
        for d in range(1, MAX_DIAS_P+1):
            cell = ws_p.cell(1, BASE_COL_P+d-1)
            cell.value     = d
            cell.font      = Font(bold=True, size=7, color="FFFFFF", name="Arial")
            cell.fill      = PatternFill("solid", fgColor="004D40")
            cell.alignment = Alignment(horizontal="center")
            ws_p.column_dimensions[get_column_letter(BASE_COL_P+d-1)].width = 1.8

        # Colores rotativos para cada PT
        PT_COLORS = ["1565C0","2E7D32","E65100","6A1B9A","00695C",
                     "AD1457","0277BD","558B2F","4E342E","37474F"]

        try:
            ini_idx_p = hdrs_p.index("Dia_Inicio") + 1
            fin_idx_p = hdrs_p.index("Dia_Fin")    + 1
        except: ini_idx_p = fin_idx_p = None

        for ri, r in enumerate(range(2, ws_p.max_row+1), 0):
            bar_color = PT_COLORS[ri % len(PT_COLORS)]
            row_fill  = PatternFill("solid", fgColor="F1F8E9" if ri % 2 == 0 else "E8F5E9")

            for c in range(1, DATA_COLS_P+1):
                cell = ws_p.cell(r, c)
                cell.fill   = row_fill
                cell.font   = Font(size=10, bold=(c<=2), name="Arial", color="212121")
                cell.border = brd
                cell.alignment = Alignment(
                    horizontal="right" if c >= 3 else "left",
                    vertical="center")

            try:
                d_ini = int(float(ws_p.cell(r, ini_idx_p).value or 0)) if ini_idx_p else 0
                d_fin = int(float(ws_p.cell(r, fin_idx_p).value or 0)) if fin_idx_p else 0
            except: d_ini = d_fin = 0

            for d in range(1, MAX_DIAS_P+1):
                cell = ws_p.cell(r, BASE_COL_P+d-1)
                if d_ini <= d <= d_fin:
                    cell.fill = PatternFill("solid", fgColor=bar_color)
                else:
                    cell.fill = PatternFill("solid", fgColor="FAFAFA")
                cell.border = Border(
                    left=Side(style="thin",color="E0E0E0"),
                    right=Side(style="thin",color="E0E0E0"))

            ws_p.row_dimensions[r].height = 22

        # Agregar etiqueta del día de fin sobre la barra (en la celda Dia_Fin+1)
        for r in range(2, ws_p.max_row+1):
            try:
                d_fin = int(float(ws_p.cell(r, fin_idx_p).value or 0)) if fin_idx_p else 0
            except: d_fin = 0
            if d_fin > 0 and d_fin < MAX_DIAS_P:
                label_cell = ws_p.cell(r, BASE_COL_P + d_fin)
                label_cell.value = f"D{d_fin}"
                label_cell.font  = Font(size=7, bold=True, color="FFFFFF", name="Arial")
                label_cell.alignment = Alignment(horizontal="center")

        # Anchos columnas de datos
        col_w_p = {"Codigo":14,"Descripcion":45,"Cantidad":12,
                   "Duracion_dias":12,"Dia_Inicio":10,"Dia_Fin":10}
        for ci, h in enumerate(hdrs_p, 1):
            ws_p.column_dimensions[get_column_letter(ci)].width = col_w_p.get(str(h), 12)
        ws_p.freeze_panes = "C2"

    # ── Formato especial hoja Resumen ────────────────────────────────────────────
    if "0_Resumen" in wb2.sheetnames:
        ws_r = wb2["0_Resumen"]
        hdrs_r = [ws_r.cell(1,c).value for c in range(1, ws_r.max_column+1)]
        # Encabezado
        for c in range(1, ws_r.max_column+1):
            cell = ws_r.cell(1,c)
            cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Arial")
            cell.fill      = PatternFill("solid", fgColor="880E4F")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = brd
        ws_r.row_dimensions[1].height = 32

        puede_idx = (hdrs_r.index("Se_puede_hacer")+1) if "Se_puede_hacer" in hdrs_r else None
        for r in range(2, ws_r.max_row+1):
            puede_val = ws_r.cell(r, puede_idx).value if puede_idx else ""
            if puede_val == "Sí":
                row_fill = PatternFill("solid", fgColor="E8F5E9")
                txt_color = "1B5E20"
            elif puede_val == "Parcial":
                row_fill = PatternFill("solid", fgColor="FFF9C4")
                txt_color = "F57F17"
            else:
                row_fill = PatternFill("solid", fgColor="FFEBEE")
                txt_color = "B71C1C"

            for c in range(1, ws_r.max_column+1):
                cell = ws_r.cell(r,c)
                cell.fill   = row_fill
                cell.border = brd
                cell.font   = Font(size=10, name="Arial",
                                   bold=(c==puede_idx),
                                   color=txt_color if c==puede_idx else "212121")
                cell.alignment = Alignment(horizontal="center" if c in {1,3,4,5,6} else "left",
                                           vertical="center")

        # Anchos
        col_w = {"Codigo":14,"Descripcion":45,"Cantidad":12,
                 "Se_puede_hacer":14,"Para_cuando_dia":14,"Cuanto":12,"Limitante":35}
        for ci, h in enumerate(hdrs_r, 1):
            ws_r.column_dimensions[get_column_letter(ci)].width = col_w.get(str(h), 14)
        ws_r.row_dimensions[1].height = 32
        for r in range(2, ws_r.max_row+1):
            ws_r.row_dimensions[r].height = 22
        ws_r.freeze_panes = "A2"

    # ── Formato especial hoja Gantt ──────────────────────────────────────────────
    if "5_Gantt" in wb2.sheetnames:
        ws_g = wb2["5_Gantt"]
        hdrs_g = [ws_g.cell(1,c).value for c in range(1, ws_g.max_column+1)]

        # Colores por puesto de trabajo
        PUESTO_COLORS = [
            "1565C0","2E7D32","E65100","6A1B9A","00695C",
            "AD1457","0277BD","558B2F","4E342E","37474F",
        ]
        puestos_unicos = []
        for r in range(2, ws_g.max_row+1):
            ft_idx = hdrs_g.index("Fila_Tipo")+1 if "Fila_Tipo" in hdrs_g else None
            pu_idx = hdrs_g.index("Puesto")+1 if "Puesto" in hdrs_g else None
            if ft_idx and ws_g.cell(r, ft_idx).value == "SEMI":
                p = ws_g.cell(r, pu_idx).value if pu_idx else ""
                if p and p not in puestos_unicos:
                    puestos_unicos.append(p)
        color_map = {p: PUESTO_COLORS[i % len(PUESTO_COLORS)] for i, p in enumerate(puestos_unicos)}

        # Columnas de datos (antes de las barras)
        DATA_COLS = len(hdrs_g)
        MAX_DIAS = 90
        BASE_COL = DATA_COLS + 2  # +1 espacio vacío

        # Encabezados de días
        ws_g.cell(1, DATA_COLS+1).value = ""
        for d in range(1, MAX_DIAS+1):
            cell = ws_g.cell(1, BASE_COL+d-1)
            cell.value = d
            cell.font      = Font(bold=True, size=7, color="FFFFFF", name="Arial")
            cell.fill      = PatternFill("solid", fgColor="37474F")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_g.column_dimensions[get_column_letter(BASE_COL+d-1)].width = 1.8

        ft_idx  = hdrs_g.index("Fila_Tipo")+1   if "Fila_Tipo"  in hdrs_g else None
        pu_idx  = hdrs_g.index("Puesto")+1       if "Puesto"     in hdrs_g else None
        ini_idx = hdrs_g.index("Dia_Inicio")+1   if "Dia_Inicio" in hdrs_g else None
        fin_idx = hdrs_g.index("Dia_Fin")+1      if "Dia_Fin"    in hdrs_g else None
        lt_idx  = hdrs_g.index("LT_dias")+1      if "LT_dias"    in hdrs_g else None
        flt_idx = hdrs_g.index("Falta_Mat")+1    if "Falta_Mat"  in hdrs_g else None

        for r in range(2, ws_g.max_row+1):
            fila_tipo = ws_g.cell(r, ft_idx).value if ft_idx else ""
            puesto    = ws_g.cell(r, pu_idx).value if pu_idx else ""
            bar_color = color_map.get(str(puesto), "90A4AE")

            if fila_tipo == "PT_HDR":
                # Encabezado del PT: fila azul marino
                for c in range(1, DATA_COLS+1):
                    ws_g.cell(r,c).fill = PatternFill("solid", fgColor="1F4E79")
                    ws_g.cell(r,c).font = Font(size=10, bold=True, name="Arial", color="FFFFFF")
                    ws_g.cell(r,c).border = brd
                for d in range(1, MAX_DIAS+1):
                    ws_g.cell(r, BASE_COL+d-1).fill = PatternFill("solid", fgColor="1F4E79")
                continue

            elif fila_tipo == "PT":
                # Fila PT ensamble: fondo verde oscuro
                for c in range(1, DATA_COLS+1):
                    ws_g.cell(r,c).fill = PatternFill("solid", fgColor="1B5E20")
                    ws_g.cell(r,c).font = Font(size=9, bold=True, name="Arial", color="FFFFFF")
                    ws_g.cell(r,c).border = brd
                try:
                    d_ini = int(float(ws_g.cell(r, ini_idx).value or 0))
                    d_fin = int(float(ws_g.cell(r, fin_idx).value or 0))
                except: d_ini = d_fin = 0
                for d in range(1, MAX_DIAS+1):
                    cell = ws_g.cell(r, BASE_COL+d-1)
                    cell.fill = PatternFill("solid", fgColor="2E7D32" if d_ini <= d <= d_fin else "1B5E20")
                    cell.border = Border(left=Side(style="thin",color="145A1F"),right=Side(style="thin",color="145A1F"))
                continue

            elif fila_tipo == "SEMI":
                # Fila semiterminado: fondo suave del color del puesto
                row_fill = PatternFill("solid", fgColor="E3F2FD")
                for c in range(1, DATA_COLS+1):
                    ws_g.cell(r,c).fill = row_fill
                    ws_g.cell(r,c).font = Font(size=9, bold=True, name="Arial", color="0D47A1")
                    ws_g.cell(r,c).border = brd

                # Dibujar barra Gantt
                try:
                    d_ini = int(float(ws_g.cell(r, ini_idx).value or 0))
                    d_fin = int(float(ws_g.cell(r, fin_idx).value or 0))
                except: d_ini = d_fin = 0

                for d in range(1, MAX_DIAS+1):
                    cell = ws_g.cell(r, BASE_COL+d-1)
                    if d_ini <= d <= d_fin:
                        cell.fill = PatternFill("solid", fgColor=bar_color)
                    else:
                        cell.fill = PatternFill("solid", fgColor="F5F5F5")
                    cell.border = Border(
                        left=Side(style="thin",color="E0E0E0"),
                        right=Side(style="thin",color="E0E0E0"))

            elif fila_tipo in ("MAT","MAT_PT"):
                # Fila material: fondo blanco/amarillo si falta
                falta_val = ws_g.cell(r, flt_idx).value if flt_idx else ""
                has_falta = falta_val not in (None, "", 0)
                row_fill  = PatternFill("solid", fgColor="FFF9C4" if has_falta else "FAFAFA")
                for c in range(1, DATA_COLS+1):
                    ws_g.cell(r,c).fill = row_fill
                    ws_g.cell(r,c).font = Font(size=8, name="Arial",
                                               color="B71C1C" if has_falta else "546E7A")
                    ws_g.cell(r,c).border = brd

                # Barra de LT (línea punteada en gris)
                lt_val = ws_g.cell(r, lt_idx).value if lt_idx else ""
                try: lt_dias = int(float(lt_val))
                except: lt_dias = 0

                for d in range(1, MAX_DIAS+1):
                    cell = ws_g.cell(r, BASE_COL+d-1)
                    if 1 <= d <= lt_dias:
                        cell.fill = PatternFill("solid", fgColor="ECEFF1")
                        cell.font = Font(size=7, color="90A4AE")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FAFAFA")
                    cell.border = Border(
                        left=Side(style="thin",color="E0E0E0"),
                        right=Side(style="thin",color="E0E0E0"))

        # Anchos columnas de datos
        col_widths = {"Fila_Tipo":0,"PT_Raiz":12,"Nivel_BOM":7,"Puesto":12,
                      "Cod_Componente":18,"Desc_Componente":42,"Und":7,
                      "Qty_A_Producir":12,"Stock_Semi":12,"Duracion_dias":10,
                      "Dia_Inicio":8,"Dia_Fin":8,"LT_dias":8,"Stock_Mat":10,"Falta_Mat":10}
        for ci, h in enumerate(hdrs_g, 1):
            w = col_widths.get(str(h), 10)
            ws_g.column_dimensions[get_column_letter(ci)].width = w
            if w == 0:
                ws_g.column_dimensions[get_column_letter(ci)].hidden = True

        ws_g.freeze_panes = "F2"
        ws_g.row_dimensions[1].height = 20

    out = io.BytesIO()
    wb2.save(out); out.seek(0)
    return dcc.send_bytes(out.read(), "BOM_Explosion.xlsx")


server = app.server

if __name__ == "__main__":
    print("Iniciando en http://127.0.0.1:8050")
    app.run(debug=False, port=8050)
